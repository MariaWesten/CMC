from __future__ import annotations

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter


BASE = Path("/Users/mariaake/Documents/CMC/outputs/time_tracking_taetigkeitsschein")
BOOK = BASE / "CMC_Time_Tracking_Taetigkeitsschein_Juni_2026.xlsx"
BACKUP = BASE / "CMC_Time_Tracking_Taetigkeitsschein_Juni_2026.backup-vor-kw28-ueberstunden.xlsx"
MAX_ROW = 214
DATA_START = 13
KW28_START = datetime(2026, 7, 6)
KW28_END = datetime(2026, 7, 12)
WEEKLY_BUDGET = 20.0


POLISHED = {
    50: "Studie 2027: Debrief mit Simone und Jule zum bisherigen Teststand; Sammlung offener Punkte und Einordnung der nächsten Arbeitsschritte.",
    51: "Studie 2027: Erstellung eines neuen Testpakets auf Basis des Debriefings sowie Weiterentwicklung des JSON-Builder-Tools für Jule und Simone, inklusive Prüfchecks und E-Mail-Output.",
    52: "Studie 2027: Kommunikation der nächsten Schritte an Jule und Simone, Auswertung der Testlaufzeiten sowie Aufbereitung eines Status-Updates zur Studie 2027.",
    53: "Studie 2027: Weiterentwicklung des JSON-Builder-Tools einschließlich Fehlerbehebung, zusätzlicher Prüfungen und optischer Optimierung; weitere Skriptoptimierung und Vorbereitung einer neuen Testrunde.",
    54: "Studie 2027: Auswertung des JSON-Builder- und Bildertests, Einarbeitung weiterer Anmerkungen sowie Zusammenstellung eines Testpakets für Simone.",
    55: "Studie 2027: Abstimmung mit Robert sowie Zusammenfassung relevanter Punkte und nächster Schritte.",
    56: "Studie 2027: Debrief mit Simone zum letzten Testlauf, Klärung offener Fragen und Ableitung weiterer Anpassungsbedarfe.",
    57: "Studie 2027: Einarbeitung in die Sales-Präsentation und Beginn der Erstellung eines Briefings für die weitere Ausarbeitung.",
    58: "Studie 2027: Überarbeitung des JSON-Builders anhand von Simones Feedback sowie Vorbereitung einer erneuten Testrunde.",
    59: "Studie 2027: Fertigstellung des Sales-Präsentationsbriefings unter Einbindung des Inputs von Maria und Lisa.",
    60: "Studie 2027: Abstimmung mit Simone zur Durchführung des Ende-zu-Ende-Tests sowie Debrief zum letzten Testlauf.",
    61: "Studie 2027: Überarbeitung des JSON-Builders anhand von Simones Feedback, Vorbereitung einer erneuten Testrunde und Aufbau eines KI-Text-Builders.",
    62: "Studie 2027: Review des aktuellen Stands der Sales-Präsentation und Weitergabe strukturierter Rückmeldungen an Cathleen.",
    63: "Studie 2027: Status-Update zur KI-Automatisierung und Einordnung der nächsten technischen Schritte.",
    64: "Studie 2027: Weekly Call mit Deutsche Post inklusive Vor- und Nachbereitung sowie Abstimmung mit Cathleen.",
}


def clone_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def german_date(dt: datetime) -> str:
    return f"{dt.day}. Juli {dt.year}"


def fmt_hours(value: float) -> str:
    return f"{value:.2f}".replace(".", ",")


def clear_sheet(ws):
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell._style = copy(ws["A1"]._style)


def style_range(ws, cell_range, fill=None, font=None, border=None, alignment=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def main():
    if not BACKUP.exists():
        shutil.copy2(BOOK, BACKUP)

    wb = load_workbook(BOOK)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    ws = wb["Datenbasis"]
    settings = wb["Einstellungen"]
    settings["B15"] = WEEKLY_BUDGET
    settings["B15"].number_format = "0.00"

    ws["K4"] = "Offen/Überstunden (Std.)"
    ws["G5"] = "=WEEKNUM(TODAY(),21)"
    ws["K5"] = '=IF(OR(I5="",J5=""),"",J5-I5)'
    ws["G6"] = "Anzeige aktualisiert sich beim Öffnen/Neuberechnen der Datei; KW-Budget steht in Einstellungen!B15. Negative Werte in K5 sind Überstunden."

    headers = {
        12: "Abrechenbar (Std.)",
        13: "Überstunden (Std.)",
        14: "KW-Gesamt (Std.)",
        15: "KW-Offen/Über (Std.)",
    }
    for col, header in headers.items():
        clone_style(ws["K12"], ws.cell(DATA_START - 1, col))
        ws.cell(DATA_START - 1, col).value = header
        ws.cell(DATA_START - 1, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(DATA_START, MAX_ROW + 1):
        ws.cell(row, 8).value = f'=IF(A{row}="","",WEEKNUM(A{row},21))'
        ws.cell(row, 7).value = f'=IF(A{row}="","",L{row}*F{row})'
        ws.cell(row, 12).value = (
            f'=IF(A{row}="","",IF(Einstellungen!$B$15="","",'
            f'MAX(0,MIN(E{row},Einstellungen!$B$15-(SUMIFS($E$13:$E{row},$H$13:$H{row},H{row})-E{row})))))'
        )
        ws.cell(row, 13).value = f'=IF(A{row}="","",IF(L{row}="","",E{row}-L{row}))'
        ws.cell(row, 14).value = f'=IF(A{row}="","",SUMIFS($E$13:$E$214,$H$13:$H$214,H{row}))'
        ws.cell(row, 15).value = f'=IF(A{row}="","",IF(Einstellungen!$B$15="","",Einstellungen!$B$15-N{row}))'
        for col in range(12, 16):
            clone_style(ws.cell(row, 5), ws.cell(row, col))
            ws.cell(row, col).number_format = "0.00"

    for row, text in POLISHED.items():
        ws.cell(row, 4).value = text
        ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col, width in {"L": 17, "M": 17, "N": 16, "O": 18}.items():
        ws.column_dimensions[col].width = width

    # Budget overview: show total, billable and overtime hours.
    budget = wb["Budgetübersicht"]
    budget_headers_month = ["Monat", "Gebucht (Std.)", "Abrechenbar (Std.)", "Überstunden (Std.)", "Budget (Std.)", "Offen (Std.)"]
    budget_headers_week = ["KW", "Gebucht (Std.)", "Abrechenbar (Std.)", "Überstunden (Std.)", "Budget (Std.)", "Offen (Std.)"]
    for idx, value in enumerate(budget_headers_month, start=1):
        budget.cell(7, idx).value = value
        clone_style(budget["A7"], budget.cell(7, idx))
    for row in range(8, 20):
        budget.cell(row, 1).value = f"=DATE(2026,{row-7},1)"
        budget.cell(row, 2).value = f'=SUMIFS(Datenbasis!$E$13:$E$214,Datenbasis!$A$13:$A$214,">="&A{row},Datenbasis!$A$13:$A$214,"<"&EDATE(A{row},1))'
        budget.cell(row, 3).value = f'=SUMIFS(Datenbasis!$L$13:$L$214,Datenbasis!$A$13:$A$214,">="&A{row},Datenbasis!$A$13:$A$214,"<"&EDATE(A{row},1))'
        budget.cell(row, 4).value = f'=SUMIFS(Datenbasis!$M$13:$M$214,Datenbasis!$A$13:$A$214,">="&A{row},Datenbasis!$A$13:$A$214,"<"&EDATE(A{row},1))'
        budget.cell(row, 5).value = '=IF(Einstellungen!$B$14="","",Einstellungen!$B$14)'
        budget.cell(row, 6).value = f'=IF(E{row}="","",MAX(0,E{row}-B{row}))'
        for col in range(1, 7):
            clone_style(budget.cell(8, min(col, 4)), budget.cell(row, col))
        budget.cell(row, 1).number_format = "[$-de-DE]MMMM YYYY"
        for col in range(2, 7):
            budget.cell(row, col).number_format = "0.00"

    for idx, value in enumerate(budget_headers_week, start=1):
        budget.cell(22, idx).value = value
        clone_style(budget["A22"], budget.cell(22, idx))
    for row in range(23, 76):
        kw = row - 22
        budget.cell(row, 1).value = kw
        budget.cell(row, 2).value = f'=IF(SUMIFS(Datenbasis!$E$13:$E$214,Datenbasis!$H$13:$H$214,A{row})=0,"",SUMIFS(Datenbasis!$E$13:$E$214,Datenbasis!$H$13:$H$214,A{row}))'
        budget.cell(row, 3).value = f'=IF(B{row}="","",SUMIFS(Datenbasis!$L$13:$L$214,Datenbasis!$H$13:$H$214,A{row}))'
        budget.cell(row, 4).value = f'=IF(B{row}="","",SUMIFS(Datenbasis!$M$13:$M$214,Datenbasis!$H$13:$H$214,A{row}))'
        budget.cell(row, 5).value = f'=IF(OR(B{row}="",Einstellungen!$B$15=""),"",Einstellungen!$B$15)'
        budget.cell(row, 6).value = f'=IF(E{row}="","",MAX(0,E{row}-B{row}))'
        for col in range(1, 7):
            clone_style(budget.cell(23, min(col, 4)), budget.cell(row, col))
        for col in range(2, 7):
            budget.cell(row, col).number_format = "0.00"

    for col, width in {"A": 18, "B": 16, "C": 18, "D": 18, "E": 16, "F": 16}.items():
        budget.column_dimensions[col].width = width

    # Build current-week screenshot view.
    z = wb["Zeiterfassung"]
    clear_sheet(z)
    z.sheet_view.showGridLines = False
    z.freeze_panes = "A2"
    z.page_setup.orientation = "landscape"
    z.page_setup.fitToWidth = 1
    z.page_setup.fitToHeight = 0
    z.sheet_properties.pageSetUpPr.fitToPage = True
    z.page_margins = PageMargins(left=0.2, right=0.2, top=0.25, bottom=0.25, header=0.1, footer=0.1)
    z.print_options.horizontalCentered = True
    z.column_dimensions["A"].width = 15
    z.column_dimensions["B"].width = 70
    z.column_dimensions["C"].width = 14
    z.column_dimensions["D"].width = 14
    z.column_dimensions["E"].width = 14

    dark = "666666"
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    section_fill = PatternFill("solid", fgColor="E8F1FD")
    overtime_fill = PatternFill("solid", fgColor="FFF2CC")
    note_fill = PatternFill("solid", fgColor="F7F7F7")
    header_font = Font(name="Aptos", bold=True, size=14)
    body_font = Font(name="Aptos", size=9)
    bold_font = Font(name="Aptos", bold=True, size=9)
    medium = Side(style="medium", color=dark)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    summary_border = Border(top=medium, bottom=medium, left=thin, right=thin)

    headers = ["Datum", "Aufgabe", "Arbeitszeit\n(Stunden)", "Abrechn.\n(Std.)", "Überstd.\n(Std.)"]
    for col, value in enumerate(headers, start=1):
        cell = z.cell(1, col)
        cell.value = value
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=medium, right=medium, top=medium, bottom=medium)
    z.row_dimensions[1].height = 44

    rows = []
    cumulative = 0.0
    for row in range(DATA_START, MAX_ROW + 1):
        date = ws.cell(row, 1).value
        if isinstance(date, datetime) and KW28_START <= date <= KW28_END and ws.cell(row, 5).value:
            hours = float(ws.cell(row, 5).value)
            billable = max(0.0, min(hours, WEEKLY_BUDGET - cumulative))
            overtime = max(0.0, hours - billable)
            cumulative += hours
            rows.append((date, ws.cell(row, 4).value, hours, billable, overtime))

    total = sum(r[2] for r in rows)
    billable_total = sum(r[3] for r in rows)
    overtime_total = sum(r[4] for r in rows)

    z.merge_cells("A2:B2")
    z["A2"] = "KW 28 (06.07.–12.07.2026)"
    z["C2"] = f"{fmt_hours(total)} Std."
    z["D2"] = f"{fmt_hours(billable_total)} Std."
    z["E2"] = f"{fmt_hours(overtime_total)} Std."
    style_range(z, "A2:E2", fill=section_fill, font=bold_font, border=summary_border, alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
    z["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    z.row_dimensions[2].height = 30

    out_row = 3
    for date, desc, hours, billable, overtime in rows:
        z.cell(out_row, 1).value = german_date(date)
        z.cell(out_row, 2).value = desc
        z.cell(out_row, 3).value = fmt_hours(hours)
        z.cell(out_row, 4).value = fmt_hours(billable)
        z.cell(out_row, 5).value = fmt_hours(overtime)
        for col in range(1, 6):
            cell = z.cell(out_row, col)
            cell.font = body_font if col != 1 else bold_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if col in (1, 2) else "right", vertical="center", wrap_text=True)
            if col == 5 and overtime > 0:
                cell.fill = overtime_fill
        z.row_dimensions[out_row].height = 49
        out_row += 1

    z.cell(out_row, 2).value = "Summe Stunden KW 28"
    z.cell(out_row, 3).value = fmt_hours(total)
    z.cell(out_row, 4).value = fmt_hours(billable_total)
    z.cell(out_row, 5).value = fmt_hours(overtime_total)
    style_range(z, f"A{out_row}:E{out_row}", fill=section_fill, font=bold_font, border=summary_border, alignment=Alignment(horizontal="right", vertical="center", wrap_text=True))
    z.cell(out_row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    z.row_dimensions[out_row].height = 28

    note_row = out_row + 1
    z.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    z.cell(note_row, 1).value = "Hinweis: Stunden oberhalb des KW-Budgets von 20,00 Std. sind als Überstunden ausgewiesen und nicht abrechenbar."
    style_range(z, f"A{note_row}:E{note_row}", fill=note_fill, font=Font(name="Aptos", italic=True, size=10), border=border, alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    z.row_dimensions[note_row].height = 28
    z.print_area = f"A1:E{note_row}"

    wb.save(BOOK)
    print(f"saved={BOOK}")
    print(f"backup={BACKUP}")
    print(f"kw28_total={total:.2f}")
    print(f"kw28_billable={billable_total:.2f}")
    print(f"kw28_overtime={overtime_total:.2f}")


if __name__ == "__main__":
    main()
